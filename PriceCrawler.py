#!/usr/bin/env python3
# coding=utf-8
import json
import logging
import re
import time
from json import decoder
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException, StaleElementReferenceException
from selenium.webdriver import DesiredCapabilities
from selenium.webdriver.chrome.options import Options

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import mysql.connector
from sqlalchemy import create_engine
from CONFIG import MYSQL_HOST, MYSQL_USER, MYSQL_PASSWORD, MYSQL_PORT, MYSQL_DATABASE, MYSQL_CHARSET

# import schedule
import datetime

class Crawler(object):

    def __init__(self, proxy=None):
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--ignore-ssl-errors')
        prefs = {"profile.managed_default_content_settings.images": 2}
        chrome_options.add_experimental_option("prefs", prefs)
        if proxy:
            proxy_address = proxy['https']
            chrome_options.add_argument('--proxy-server=%s' % proxy_address)
            logging.info('Chrome using proxy: %s', proxy['https'])

        self.chrome = webdriver.Chrome(options=chrome_options) #, desired_capabilities=caps
        # jd sometimes load google pic takes much time
        self.chrome.set_page_load_timeout(30)
        # set timeout for script
        self.chrome.set_script_timeout(30)

        self.found=False

    def close(self):
        self.chrome.quit()

    def get_jd_item(self, url):
        item_info_dict = {"name": None, "price": None, "plus_price": None, "subtitle": None}
        # url = 'https://item.jd.com/' + item_id + '.html'
        url = url
        try:
            self.chrome.get(url)
            logging.info('Crawl: {}'.format(url))
            # 共8秒
            retry = 5
            while retry:
                try:
                    element = self.chrome.find_element("xpath","//*[@class='p-price']/span[2]").text
                    if element:
                        logging.info("爬取价格数据")
                        logging.info('Found price element: {}'.format(element))
                        # global found
                        self.found = True
                        break
                    else:
                        logging.info("价格元素出现，价格未出现重试2秒")
                        time.sleep(2)
                        retry -= 1
                        if retry == 1: break
                except NoSuchElementException:
                    logging.info("价格元素未出现")
                    self.found = False
                    time.sleep(2)
                    retry -= 1
                    if retry == 1: break
        except TimeoutException as e:
            logging.warning('Crawl failure: {}'.format(e.msg))
            return item_info_dict

        # 提取商品名称
        try:
            name = self.chrome.find_element("xpath","//*[@class='sku-name']").text
            item_info_dict['name'] = name
        except AttributeError as e:
            logging.warning('Crawl name failure: {}'.format(e))
        except NoSuchElementException:
            try:
                # 加油卡(如200117841739）需要改为提取name
                name = self.chrome.find_element("xpath","//*[@class='name']").text
                item_info_dict['name'] = name
            except NoSuchElementException as e:
                logging.warning('Crawl name failure: {}'.format(e.msg))

        # 提取商品价格
        try:
            # global found
            if self.found:
                price = self.chrome.find_element("xpath","//*[@class='p-price']").text
                if price:
                    price_xpath = re.findall(r'-?\d+\.?\d*e?-?\d*?', price)
                    if price_xpath:  # 若能提取到值
                        item_info_dict['price'] = price_xpath[0]  # 提取浮点数
            # elif self.chrome.find_element("xpath","//*[@class='itemover-tip']").text:
            else:
                item_info_dict['price'] ='商品已下架'
        except AttributeError as e:
            logging.warning('Crawl price failure: {}'.format(e.msg))
        except NoSuchElementException as e:
            logging.warning('Crawl price failure: {}'.format(e.msg))
            return item_info_dict

        logging.info('Crawl SUCCESS: {}'.format(item_info_dict))
        return item_info_dict
    

if __name__ == '__main__':
    logging.basicConfig(format="%(asctime)s | %(levelname)s | %(filename)s %(lineno)s | %(message)s",
                        datefmt="%Y-%m-%d %H:%M:%S",
                        level=logging.INFO)

    c = Crawler()

    def update_prices():
        # logging.debug(c.get_jd_item('100023130207'))
        excel_file = "prices.xlsx"
        wb = openpyxl.load_workbook(excel_file)
        for sheet in wb.sheetnames:
            current_sheet = wb[sheet]

            current_date = datetime.date.today().strftime("%Y-%m-%d")

            if current_sheet.cell(row = 1, column=current_sheet.max_column).value != current_date:
                column_letter = get_column_letter(current_sheet.max_column + 1)
                current_sheet[column_letter + '1'] = current_date
                current_sheet[column_letter + '1'].font = Font(bold=True)
            
            product_row = None
            for row in current_sheet.iter_rows(min_row=2, max_row=current_sheet.max_row, min_col=4, max_col=4):
                url = row[0].value
                product_row = row[0].row
                price = c.get_jd_item(url)['price']
                current_sheet.cell(row=product_row, column=current_sheet.max_column, value=price)

        wb.save(excel_file)
        wb.close()
    
    update_prices()

    def update_prices_SQL():
        # 连接到数据库
        conn = mysql.connector.connect(
            host=MYSQL_HOST,
            port=MYSQL_PORT,
            user=MYSQL_USER,
            password=MYSQL_PASSWORD,
            database=MYSQL_DATABASE
        )
        cursor = conn.cursor()

        # 获取表名
        cursor.execute("SHOW TABLES")
        tables = cursor.fetchall()

        current_date = 'T' + datetime.date.today().strftime("%Y%m%d")

        for table in tables:
            table_name = table[0]
            cursor.execute(f" ALTER TABLE {table_name} ADD {current_date} VARCHAR(255)")
           
            # 查询商品名称和在售链接
            cursor.execute(f"SELECT num, url FROM {table_name}")
            products = cursor.fetchall()

            for product in products:
                num = product[0]
                url = product[1]
                price = c.get_jd_item(url)['price']  # 从URL中获取价格

                # 更新数据库，将价格数据插入到新字段中
                update_query = f"UPDATE {table_name} SET {current_date} = '{price}' WHERE num = {num}"
                cursor.execute(update_query)
                conn.commit()

        # 关闭数据库连接
        cursor.close()
        conn.close()

    # update_prices_SQL()

    c.close()
